from __future__ import annotations

import os
import re
import tempfile
from datetime import date, datetime
from functools import wraps
from pathlib import Path
from typing import Any

from flask import (
    Flask,
    Response,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
LOCAL_DB_PATH = BASE_DIR / "sim_bonus.db"


def normalize_database_url(raw_url: str | None) -> str:
    """Railway Postgres gives DATABASE_URL. Locally we use SQLite."""
    if not raw_url:
        return f"sqlite:///{LOCAL_DB_PATH}"
    if raw_url.startswith("postgres://"):
        return raw_url.replace("postgres://", "postgresql+psycopg://", 1)
    if raw_url.startswith("postgresql://") and "+" not in raw_url.split("://", 1)[0]:
        return raw_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return raw_url


DATABASE_URL = normalize_database_url(os.getenv("DATABASE_URL"))
CONNECT_ARGS = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine: Engine = create_engine(
    DATABASE_URL,
    future=True,
    pool_pre_ping=True,
    connect_args=CONNECT_ARGS,
)


def create_app() -> Flask:
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret-key-change-in-production")

    init_db()

    @app.context_processor
    def inject_now() -> dict[str, Any]:
        return {"now": datetime.now(), "current_year": datetime.now().year}

    @app.route("/health")
    def health() -> str:
        return "ok"

    @app.route("/sw.js")
    def service_worker() -> Response:
        response = send_from_directory(app.static_folder, "sw.js", mimetype="application/javascript")
        response.headers["Cache-Control"] = "no-cache"
        return response

    @app.route("/")
    @login_required
    def dashboard() -> str:
        today = date.today().isoformat()
        month = request.args.get("month") or date.today().strftime("%Y-%m")
        operator_id = parse_int(request.args.get("operator_id"))
        operator_where = " AND t.operator_id = :operator_id" if operator_id else ""
        params = {"month": month, "today": today, "operator_id": operator_id}
        with engine.connect() as conn:
            operators = fetch_all(conn, "SELECT * FROM operators ORDER BY is_active DESC, name")
            total_today = scalar(
                conn,
                f"""
                SELECT COALESCE(SUM(s.quantity), 0) AS total
                FROM sales s
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE s.sale_date = :today {operator_where}
                """,
                params,
            )
            total_month = scalar(
                conn,
                f"""
                SELECT COALESCE(SUM(s.quantity), 0) AS total
                FROM sales s
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
                """,
                params,
            )
            bonus_month = scalar(
                conn,
                f"""
                SELECT COALESCE(SUM(s.quantity * t.bonus_per_item), 0) AS total
                FROM sales s
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
                """,
                params,
            )
            top_employees = fetch_all(
                conn,
                f"""
                SELECT e.full_name, COALESCE(SUM(s.quantity), 0) AS qty,
                       COALESCE(SUM(s.quantity * t.bonus_per_item), 0) AS bonus
                FROM sales s
                JOIN employees e ON e.id = s.employee_id
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
                GROUP BY e.id, e.full_name
                ORDER BY qty DESC, e.full_name
                LIMIT 10
                """,
                params,
            )
            top_tariffs = fetch_all(
                conn,
                f"""
                SELECT t.name, o.name AS operator_name, COALESCE(SUM(s.quantity), 0) AS qty
                FROM sales s
                JOIN tariffs t ON t.id = s.tariff_id
                JOIN operators o ON o.id = t.operator_id
                WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
                GROUP BY t.id, t.name, o.id, o.name
                ORDER BY qty DESC, o.name, t.name
                LIMIT 10
                """,
                params,
            )
        return render_template(
            "dashboard.html",
            active="dashboard",
            month=month,
            operator_id=operator_id,
            operators=operators,
            total_today=total_today,
            total_month=total_month,
            bonus_month=bonus_month,
            top_employees=top_employees,
            top_tariffs=top_tariffs,
        )

    @app.route("/login", methods=["GET", "POST"])
    def login() -> str | Response:
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            with engine.connect() as conn:
                user = fetch_one(conn, "SELECT * FROM users WHERE username = :username", {"username": username})
            if user and check_password_hash(user["password_hash"], password):
                session.clear()
                session["user_id"] = user["id"]
                session["username"] = user["username"]
                flash("Хуш келибсиз!", "success")
                return redirect(url_for("dashboard"))
            flash("Логин ёки пароль нотўғри.", "danger")
        return render_template("login.html")

    @app.route("/logout")
    def logout() -> Response:
        session.clear()
        flash("Тизимдан чиқдингиз.", "info")
        return redirect(url_for("login"))

    @app.route("/password", methods=["GET", "POST"])
    @login_required
    def change_password() -> str | Response:
        if request.method == "POST":
            old_password = request.form.get("old_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if len(new_password) < 6:
                flash("Янги пароль камида 6 та белгидан иборат бўлиши керак.", "danger")
                return redirect(url_for("change_password"))
            if new_password != confirm_password:
                flash("Янги пароль такрори мос эмас.", "danger")
                return redirect(url_for("change_password"))
            with engine.begin() as conn:
                user = fetch_one(conn, "SELECT * FROM users WHERE id = :id", {"id": session["user_id"]})
                if not user or not check_password_hash(user["password_hash"], old_password):
                    flash("Эски пароль нотўғри.", "danger")
                    return redirect(url_for("change_password"))
                conn.execute(
                    text("UPDATE users SET password_hash = :hash WHERE id = :id"),
                    {"hash": generate_password_hash(new_password), "id": session["user_id"]},
                )
                add_audit(conn, "Пароль ўзгартирилди", session.get("username", "admin"))
            flash("Пароль ўзгартирилди.", "success")
            return redirect(url_for("dashboard"))
        return render_template("password.html", active="password")

    @app.route("/employees")
    @login_required
    def employees() -> str:
        with engine.connect() as conn:
            rows = fetch_all(conn, "SELECT * FROM employees ORDER BY is_active DESC, full_name")
        return render_template("employees.html", active="employees", rows=rows)

    @app.route("/employees/add", methods=["POST"])
    @login_required
    def add_employee() -> Response:
        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            flash("Ходим Ф.И.О. киритилмаган.", "danger")
            return redirect(url_for("employees"))
        try:
            with engine.begin() as conn:
                existing = fetch_one(
                    conn,
                    "SELECT id FROM employees WHERE lower(trim(full_name)) = lower(trim(:full_name))",
                    {"full_name": full_name},
                )
                if existing:
                    flash("Бу Ф.И.О. аллақачон бор.", "warning")
                    return redirect(url_for("employees"))
                conn.execute(
                    text(
                        """
                        INSERT INTO employees(id, full_name, is_active, created_at)
                        VALUES(:id, :full_name, 1, :created_at)
                        """
                    ),
                    {"id": next_id(conn, "employees"), "full_name": full_name, "created_at": current_timestamp()},
                )
                add_audit(conn, "Ходим қўшилди", full_name)
            flash("Ходим қўшилди.", "success")
        except IntegrityError as exc:
            flash(integrity_message(exc, "Бу Ф.И.О. аллақачон бор."), "warning")
        except SQLAlchemyError as exc:
            flash(f"Ходим қўшишда база хатоси: {db_error_text(exc)}", "danger")
        return redirect(url_for("employees"))

    @app.route("/employees/<int:employee_id>/edit", methods=["POST"])
    @login_required
    def edit_employee(employee_id: int) -> Response:
        full_name = request.form.get("full_name", "").strip()
        is_active = 1 if request.form.get("is_active") == "on" else 0
        if not full_name:
            flash("Ходим Ф.И.О. бўш бўлмаслиги керак.", "danger")
            return redirect(url_for("employees"))
        with engine.begin() as conn:
            try:
                conn.execute(
                    text("UPDATE employees SET full_name = :full_name, is_active = :is_active WHERE id = :id"),
                    {"full_name": full_name, "is_active": is_active, "id": employee_id},
                )
                add_audit(conn, "Ходим маълумоти ўзгартирилди", f"employee_id={employee_id}; {full_name}")
                flash("Ходим маълумоти сақланди.", "success")
            except IntegrityError:
                flash("Бу Ф.И.О. бошқа ходимда бор.", "warning")
        return redirect(url_for("employees"))

    @app.route("/operators")
    @login_required
    def operators() -> str:
        with engine.connect() as conn:
            rows = fetch_all(conn, "SELECT * FROM operators ORDER BY is_active DESC, name")
        return render_template("operators.html", active="operators", rows=rows)

    @app.route("/operators/add", methods=["POST"])
    @login_required
    def add_operator() -> Response:
        name = request.form.get("name", "").strip()
        if not name:
            flash("Оператор номи киритилмаган.", "danger")
            return redirect(url_for("operators"))
        try:
            with engine.begin() as conn:
                existing = fetch_one(
                    conn,
                    "SELECT id FROM operators WHERE lower(trim(name)) = lower(trim(:name))",
                    {"name": name},
                )
                if existing:
                    flash("Бу оператор аллақачон бор.", "warning")
                    return redirect(url_for("operators"))
                conn.execute(
                    text(
                        """
                        INSERT INTO operators(id, name, is_active, created_at)
                        VALUES(:id, :name, 1, :created_at)
                        """
                    ),
                    {"id": next_id(conn, "operators"), "name": name, "created_at": current_timestamp()},
                )
                add_audit(conn, "Оператор қўшилди", name)
            flash("Оператор қўшилди.", "success")
        except IntegrityError:
            flash("Бу оператор аллақачон бор.", "warning")
        except SQLAlchemyError as exc:
            flash(f"Оператор қўшишда база хатоси: {db_error_text(exc)}", "danger")
        return redirect(url_for("operators"))

    @app.route("/operators/<int:operator_id>/edit", methods=["POST"])
    @login_required
    def edit_operator(operator_id: int) -> Response:
        name = request.form.get("name", "").strip()
        is_active = 1 if request.form.get("is_active") == "on" else 0
        if not name:
            flash("Оператор номи бўш бўлмаслиги керак.", "danger")
            return redirect(url_for("operators"))
        try:
            with engine.begin() as conn:
                duplicate = fetch_one(
                    conn,
                    """
                    SELECT id FROM operators
                    WHERE lower(trim(name)) = lower(trim(:name)) AND id <> :id
                    """,
                    {"name": name, "id": operator_id},
                )
                if duplicate:
                    flash("Бу оператор номи бошқа операторда бор.", "warning")
                    return redirect(url_for("operators"))
                conn.execute(
                    text("UPDATE operators SET name = :name, is_active = :is_active WHERE id = :id"),
                    {"name": name, "is_active": is_active, "id": operator_id},
                )
                add_audit(conn, "Оператор ўзгартирилди", f"operator_id={operator_id}; {name}")
            flash("Оператор маълумоти сақланди.", "success")
        except SQLAlchemyError as exc:
            flash(f"Операторни сақлашда база хатоси: {db_error_text(exc)}", "danger")
        return redirect(url_for("operators"))

    @app.route("/tariffs")
    @login_required
    def tariffs() -> str:
        operator_id = parse_int(request.args.get("operator_id"))
        with engine.connect() as conn:
            operators_list = fetch_all(conn, "SELECT * FROM operators ORDER BY is_active DESC, name")
            sql = """
                SELECT t.*, o.name AS operator_name
                FROM tariffs t
                JOIN operators o ON o.id = t.operator_id
            """
            params: dict[str, Any] = {}
            if operator_id:
                sql += " WHERE t.operator_id = :operator_id"
                params["operator_id"] = operator_id
            sql += " ORDER BY o.name, t.is_active DESC, t.name"
            rows = fetch_all(conn, sql, params)
        return render_template(
            "tariffs.html",
            active="tariffs",
            rows=rows,
            operators=operators_list,
            operator_id=operator_id,
        )

    @app.route("/tariffs/add", methods=["POST"])
    @login_required
    def add_tariff() -> Response:
        operator_id = parse_int(request.form.get("operator_id"))
        name = tariff_name_display(request.form.get("name", ""))
        bonus = parse_int(request.form.get("bonus_per_item"))
        if operator_id <= 0:
            flash("Оператор танланмаган.", "danger")
            return redirect(url_for("tariffs"))
        if not name:
            flash("Тариф номи киритилмаган.", "danger")
            return redirect(url_for("tariffs"))
        if bonus < 0:
            flash("Бонус суммаси манфий бўлмаслиги керак.", "danger")
            return redirect(url_for("tariffs"))
        try:
            with engine.begin() as conn:
                existing = fetch_one(
                    conn,
                    """
                    SELECT id FROM tariffs
                    WHERE operator_id = :operator_id
                      AND lower(trim(name)) = lower(trim(:name))
                    """,
                    {"operator_id": operator_id, "name": name},
                )
                if existing:
                    flash("Бу тариф аллақачон бор.", "warning")
                    return redirect(url_for("tariffs"))
                conn.execute(
                    text(
                        """
                        INSERT INTO tariffs(id, operator_id, name, bonus_per_item, is_active, created_at)
                        VALUES(:id, :operator_id, :name, :bonus, 1, :created_at)
                        """
                    ),
                    {
                        "id": next_id(conn, "tariffs"),
                        "operator_id": operator_id,
                        "name": name,
                        "bonus": bonus,
                        "created_at": current_timestamp(),
                    },
                )
                operator = fetch_one(conn, "SELECT name FROM operators WHERE id = :id", {"id": operator_id})
                operator_name = operator["name"] if operator else str(operator_id)
                add_audit(conn, "Тариф қўшилди", f"{operator_name}; {name}; бонус={bonus}")
            flash("Тариф қўшилди.", "success")
        except IntegrityError as exc:
            flash(integrity_message(exc, "Бу тариф аллақачон бор."), "warning")
        except SQLAlchemyError as exc:
            flash(f"Тариф қўшишда база хатоси: {db_error_text(exc)}", "danger")
        return redirect(url_for("tariffs"))

    @app.route("/tariffs/<int:tariff_id>/edit", methods=["POST"])
    @login_required
    def edit_tariff(tariff_id: int) -> Response:
        operator_id = parse_int(request.form.get("operator_id"))
        name = tariff_name_display(request.form.get("name", ""))
        bonus = parse_int(request.form.get("bonus_per_item"))
        is_active = 1 if request.form.get("is_active") == "on" else 0
        if operator_id <= 0:
            flash("Оператор танланмаган.", "danger")
            return redirect(url_for("tariffs"))
        if not name:
            flash("Тариф номи бўш бўлмаслиги керак.", "danger")
            return redirect(url_for("tariffs"))
        if bonus < 0:
            flash("Бонус суммаси манфий бўлмаслиги керак.", "danger")
            return redirect(url_for("tariffs"))
        with engine.begin() as conn:
            try:
                duplicate = fetch_one(
                    conn,
                    """
                    SELECT id FROM tariffs
                    WHERE operator_id = :operator_id
                      AND lower(trim(name)) = lower(trim(:name))
                      AND id <> :id
                    """,
                    {"operator_id": operator_id, "name": name, "id": tariff_id},
                )
                if duplicate:
                    flash("Бу операторда ушбу тариф аллақачон бор.", "warning")
                    return redirect(url_for("tariffs"))
                conn.execute(
                    text(
                        """
                        UPDATE tariffs
                        SET operator_id = :operator_id,
                            name = :name,
                            bonus_per_item = :bonus,
                            is_active = :is_active
                        WHERE id = :id
                        """
                    ),
                    {
                        "operator_id": operator_id,
                        "name": name,
                        "bonus": bonus,
                        "is_active": is_active,
                        "id": tariff_id,
                    },
                )
                operator = fetch_one(conn, "SELECT name FROM operators WHERE id = :id", {"id": operator_id})
                operator_name = operator["name"] if operator else str(operator_id)
                add_audit(
                    conn,
                    "Тариф ўзгартирилди",
                    f"tariff_id={tariff_id}; {operator_name}; {name}; бонус={bonus}",
                )
                flash("Тариф сақланди.", "success")
            except IntegrityError:
                flash("Бу операторда ушбу тариф номи аллақачон бор.", "warning")
        return redirect(url_for("tariffs"))

    @app.route("/sales/new", methods=["GET", "POST"])
    @login_required
    def sale_new() -> str | Response:
        if request.method == "POST":
            sale_date = request.form.get("sale_date", "").strip()
            employee_id = parse_int(request.form.get("employee_id"))
            comment = request.form.get("comment", "").strip()
            quantities, has_negative = collect_tariff_quantities(request.form)

            error = validate_daily_sale_form(sale_date, employee_id, quantities, has_negative)
            if error:
                flash(error, "danger")
                return redirect(url_for("sale_new", sale_date=sale_date, employee_id=employee_id))
            if is_month_closed(sale_date[:7]):
                flash("Бу ой ёпилган. Маълумот киритиш ёки ўзгартириш мумкин эмас.", "danger")
                return redirect(url_for("sale_new", sale_date=sale_date, employee_id=employee_id))

            inserted, updated = save_daily_sales(
                sale_date=sale_date,
                employee_id=employee_id,
                quantities=quantities,
                comment=comment,
                created_by=session.get("username", "admin"),
            )
            if inserted or updated:
                parts = []
                if inserted:
                    parts.append(f"{inserted} та янги тариф")
                if updated:
                    parts.append(f"{updated} та аввал киритилган тариф янгиланди")
                flash("Кунлик сотув сақланди: " + ", ".join(parts) + ".", "success")
                return redirect(url_for("sale_new", sale_date=sale_date))

            flash("Сақлаш учун мос тариф топилмади.", "warning")
            return redirect(url_for("sale_new", sale_date=sale_date, employee_id=employee_id))

        employees_list, tariffs_list = get_active_refs(positive_tariffs_only=True)
        selected_date = request.args.get("sale_date", "").strip() or date.today().isoformat()
        selected_employee_id = parse_int(request.args.get("employee_id"))
        return render_template(
            "sale_form.html",
            active="sale_new",
            sale=None,
            employees=employees_list,
            tariffs=tariffs_list,
            today=selected_date,
            selected_employee_id=selected_employee_id,
        )

    @app.route("/sales/<int:sale_id>/edit", methods=["GET", "POST"])
    @login_required
    def sale_edit(sale_id: int) -> str | Response:
        with engine.connect() as conn:
            sale = fetch_one(
                conn,
                """
                SELECT s.*, t.operator_id, o.name AS operator_name
                FROM sales s
                JOIN tariffs t ON t.id = s.tariff_id
                JOIN operators o ON o.id = t.operator_id
                WHERE s.id = :id
                """,
                {"id": sale_id},
            )
        if not sale:
            flash("Сотув топилмади.", "danger")
            return redirect(url_for("sales_journal"))

        if request.method == "POST":
            sale_date = request.form.get("sale_date", "").strip()
            employee_id = parse_int(request.form.get("employee_id"))
            tariff_id = parse_int(request.form.get("tariff_id"))
            quantity = parse_int(request.form.get("quantity"))
            comment = request.form.get("comment", "").strip()
            reason = request.form.get("reason", "").strip()

            error = validate_sale_form(sale_date, employee_id, tariff_id, quantity)
            if error:
                flash(error, "danger")
                return redirect(url_for("sale_edit", sale_id=sale_id))
            if not reason:
                flash("Ўзгартириш сабаби киритилиши шарт.", "danger")
                return redirect(url_for("sale_edit", sale_id=sale_id))
            if is_month_closed(sale_date[:7]) or is_month_closed(sale["sale_date"][:7]):
                flash("Бу ой ёпилган. Маълумотни ўзгартириш мумкин эмас.", "danger")
                return redirect(url_for("sales_journal"))

            with engine.begin() as conn:
                duplicate = fetch_one(
                    conn,
                    """
                    SELECT id FROM sales
                    WHERE sale_date = :sale_date AND employee_id = :employee_id
                      AND tariff_id = :tariff_id AND id <> :id
                    """,
                    {"sale_date": sale_date, "employee_id": employee_id, "tariff_id": tariff_id, "id": sale_id},
                )
                if duplicate:
                    flash("Бу санада ушбу ходим учун бу тариф аллақачон киритилган.", "warning")
                    return redirect(url_for("sale_edit", sale_id=sale_id))
                conn.execute(
                    text(
                        """
                        UPDATE sales
                        SET sale_date = :sale_date,
                            employee_id = :employee_id,
                            tariff_id = :tariff_id,
                            quantity = :quantity,
                            comment = :comment,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = :id
                        """
                    ),
                    {
                        "sale_date": sale_date,
                        "employee_id": employee_id,
                        "tariff_id": tariff_id,
                        "quantity": quantity,
                        "comment": comment,
                        "id": sale_id,
                    },
                )
                add_audit(conn, "Сотув ўзгартирилди", f"sale_id={sale_id}; сабаб: {reason}")
            flash("Сотув ўзгартирилди.", "success")
            return redirect(url_for("sales_journal"))

        employees_list, tariffs_list = get_active_refs(include_ids=(sale["employee_id"], sale["tariff_id"]))
        return render_template(
            "sale_form.html",
            active="sales_journal",
            sale=sale,
            employees=employees_list,
            tariffs=tariffs_list,
            today=date.today().isoformat(),
        )

    @app.route("/sales/<int:sale_id>/delete", methods=["POST"])
    @login_required
    def sale_delete(sale_id: int) -> Response:
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Ўчириш сабаби киритилиши шарт.", "danger")
            return redirect(url_for("sales_journal"))
        with engine.begin() as conn:
            sale = fetch_one(conn, "SELECT * FROM sales WHERE id = :id", {"id": sale_id})
            if not sale:
                flash("Сотув топилмади.", "danger")
                return redirect(url_for("sales_journal"))
            if is_month_closed(sale["sale_date"][:7]):
                flash("Бу ой ёпилган. Маълумотни ўчириш мумкин эмас.", "danger")
                return redirect(url_for("sales_journal"))
            conn.execute(text("DELETE FROM sales WHERE id = :id"), {"id": sale_id})
            add_audit(conn, "Сотув ўчирилди", f"sale_id={sale_id}; сабаб: {reason}")
        flash("Сотув ўчирилди.", "success")
        return redirect(url_for("sales_journal"))

    @app.route("/sales")
    @login_required
    def sales_journal() -> str:
        filters = get_filters()
        rows = fetch_sales(filters)
        employees_list, tariffs_list, operators_list = get_all_refs()
        return render_template(
            "sales_journal.html",
            active="sales_journal",
            rows=rows,
            filters=filters,
            employees=employees_list,
            tariffs=tariffs_list,
            operators=operators_list,
        )

    @app.route("/report")
    @login_required
    def report() -> str:
        month = request.args.get("month") or date.today().strftime("%Y-%m")
        operator_id = parse_int(request.args.get("operator_id"))
        data = build_report(month, operator_id)
        closed = is_month_closed(month)
        with engine.connect() as conn:
            operators_list = fetch_all(conn, "SELECT * FROM operators ORDER BY is_active DESC, name")
        return render_template(
            "report.html",
            active="report",
            month=month,
            operator_id=operator_id,
            operators=operators_list,
            closed=closed,
            **data,
        )

    @app.route("/report/export")
    @login_required
    def report_export() -> Response:
        month = request.args.get("month") or date.today().strftime("%Y-%m")
        operator_id = parse_int(request.args.get("operator_id"))
        data = build_report(month, operator_id)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        file_path = Path(tmp.name)
        create_excel_report(file_path, month, data)
        suffix = f"_operator_{operator_id}" if operator_id else "_barcha_operatorlar"
        return send_file(file_path, as_attachment=True, download_name=f"bonus_hisobot_{month}{suffix}.xlsx")

    @app.route("/month/close", methods=["POST"])
    @login_required
    def month_close() -> Response:
        month = request.form.get("month", "").strip()
        operator_id = parse_int(request.form.get("operator_id"))
        if not is_valid_month(month):
            flash("Ой формати нотўғри.", "danger")
            return redirect(url_for("report", operator_id=operator_id))
        with engine.begin() as conn:
            existing = fetch_one(conn, "SELECT month FROM month_locks WHERE month = :month", {"month": month})
            if not existing:
                conn.execute(
                    text("INSERT INTO month_locks(month, closed_by, closed_at) VALUES(:month, :closed_by, CURRENT_TIMESTAMP)"),
                    {"month": month, "closed_by": session.get("username", "admin")},
                )
                add_audit(conn, "Ой ёпилди", month)
        flash(f"{month} ойи ёпилди.", "success")
        return redirect(url_for("report", month=month, operator_id=operator_id))

    @app.route("/month/open", methods=["POST"])
    @login_required
    def month_open() -> Response:
        month = request.form.get("month", "").strip()
        operator_id = parse_int(request.form.get("operator_id"))
        reason = request.form.get("reason", "").strip()
        if not reason:
            flash("Ойни қайта очиш сабаби киритилиши шарт.", "danger")
            return redirect(url_for("report", month=month, operator_id=operator_id))
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM month_locks WHERE month = :month"), {"month": month})
            add_audit(conn, "Ой қайта очилди", f"{month}; сабаб: {reason}")
        flash(f"{month} ойи қайта очилди.", "success")
        return redirect(url_for("report", month=month, operator_id=operator_id))

    @app.route("/audit")
    @login_required
    def audit() -> str:
        with engine.connect() as conn:
            rows = fetch_all(conn, "SELECT * FROM audit_logs ORDER BY created_at DESC, id DESC LIMIT 200")
        return render_template("audit.html", active="audit", rows=rows)

    return app


def fetch_one(conn, sql: str, params: dict[str, Any] | None = None):
    return conn.execute(text(sql), params or {}).mappings().first()


def fetch_all(conn, sql: str, params: dict[str, Any] | None = None) -> list[Any]:
    return list(conn.execute(text(sql), params or {}).mappings().all())


def scalar(conn, sql: str, params: dict[str, Any] | None = None) -> Any:
    return conn.execute(text(sql), params or {}).scalar() or 0


ALLOWED_ID_TABLES = {"users", "employees", "operators", "tariffs", "sales", "audit_logs"}


def current_timestamp() -> datetime:
    """Use an explicit timestamp so old Railway DB tables do not need DEFAULT created_at."""
    return datetime.now().replace(microsecond=0)


def next_id(conn, table_name: str) -> int:
    """Generate a safe id even if an old Postgres table has no identity/serial default."""
    if table_name not in ALLOWED_ID_TABLES:
        raise ValueError(f"Unsupported table name for next_id: {table_name}")
    return int(conn.execute(text(f"SELECT COALESCE(MAX(id), 0) + 1 FROM {table_name}")).scalar() or 1)


def db_error_text(exc: Exception) -> str:
    message = str(getattr(exc, "orig", exc)).replace("\n", " ").strip()
    return message[:450] if message else exc.__class__.__name__


def is_unique_error(exc: Exception) -> bool:
    message = db_error_text(exc).lower()
    unique_markers = ("unique", "duplicate key", "already exists", "уник", "takror")
    return any(marker in message for marker in unique_markers)


def integrity_message(exc: Exception, duplicate_message: str) -> str:
    if is_unique_error(exc):
        return duplicate_message
    return f"База структурасида хато: {db_error_text(exc)}"


def init_db() -> None:
    dialect = engine.dialect.name
    if dialect == "postgresql":
        id_column = "INTEGER PRIMARY KEY GENERATED BY DEFAULT AS IDENTITY"
    else:
        id_column = "INTEGER PRIMARY KEY AUTOINCREMENT"

    schema = f"""
    CREATE TABLE IF NOT EXISTS users (
        id {id_column},
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS employees (
        id {id_column},
        full_name TEXT NOT NULL UNIQUE,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS operators (
        id {id_column},
        name TEXT NOT NULL UNIQUE,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS tariffs (
        id {id_column},
        operator_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        bonus_per_item INTEGER NOT NULL DEFAULT 0,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(operator_id) REFERENCES operators(id),
        UNIQUE(operator_id, name)
    );

    CREATE TABLE IF NOT EXISTS sales (
        id {id_column},
        sale_date TEXT NOT NULL,
        employee_id INTEGER NOT NULL,
        tariff_id INTEGER NOT NULL,
        quantity INTEGER NOT NULL,
        comment TEXT,
        created_by TEXT,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP,
        FOREIGN KEY(employee_id) REFERENCES employees(id),
        FOREIGN KEY(tariff_id) REFERENCES tariffs(id),
        UNIQUE(sale_date, employee_id, tariff_id)
    );

    CREATE TABLE IF NOT EXISTS month_locks (
        month TEXT PRIMARY KEY,
        closed_by TEXT,
        closed_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS audit_logs (
        id {id_column},
        action TEXT NOT NULL,
        details TEXT,
        username TEXT,
        created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    """
    with engine.begin() as conn:
        if dialect == "postgresql":
            # Gunicorn бир нечта worker'ни бир вақтда ишга туширади.
            # Барча worker'лар бир пайтда DDL бажармаслиги учун база
            # инициализациясини PostgreSQL advisory lock билан навбатлаймиз.
            conn.execute(text("SELECT pg_advisory_xact_lock(CAST(hashtext('gsm_sim_bonus_init_db_v2') AS bigint))"))

            # Аввалги параллел старт ярим йўлда тўхтаган бўлса, identity
            # sequence қолиб, operators жадвали яратилмаган бўлиши мумкин.
            # Фақат жадвал йўқ ва sequence бор ҳолатдагина уни тозалаймиз.
            orphan_operator_sequence = conn.execute(
                text(
                    """
                    SELECT
                        to_regclass('public.operators') IS NULL AS table_missing,
                        to_regclass('public.operators_id_seq') IS NOT NULL AS sequence_exists
                    """
                )
            ).mappings().first()
            if (
                orphan_operator_sequence
                and orphan_operator_sequence["table_missing"]
                and orphan_operator_sequence["sequence_exists"]
            ):
                conn.execute(text("DROP SEQUENCE public.operators_id_seq"))

        for statement in schema.split(";"):
            statement = statement.strip()
            if statement:
                conn.execute(text(statement))

        # Агар база аввал бошқа версия/бошқа проект учун ишлатилган бўлса,
        # CREATE TABLE IF NOT EXISTS мавжуд жадвалларга янги устун қўшмайди.
        # Шунинг учун керакли устунларни хавфсиз тарзда қўшиб чиқамиз.
        ensure_schema_compatibility(conn)
        ensure_default_operators(conn)

        # Эски база қайта ишлатилганда users жадвалида бошқа фойдаланувчилар бўлиши мумкин.
        # Шунинг учун admin бор-йўқлигини умумий user_count билан эмас, username орқали текширамиз.
        admin_user = fetch_one(conn, "SELECT id, password_hash FROM users WHERE username = :username", {"username": "admin"})
        if not admin_user:
            conn.execute(
                text(
                    """
                    INSERT INTO users(id, username, password_hash, created_at)
                    VALUES(:id, :username, :password_hash, :created_at)
                    """
                ),
                {
                    "id": next_id(conn, "users"),
                    "username": "admin",
                    "password_hash": generate_password_hash("admin123"),
                    "created_at": current_timestamp(),
                },
            )
            conn.execute(
                text(
                    """
                    INSERT INTO audit_logs(id, action, details, username, created_at)
                    VALUES(:id, :action, :details, :username, :created_at)
                    """
                ),
                {
                    "id": next_id(conn, "audit_logs"),
                    "action": "Тизим яратилди",
                    "details": "Биринчи admin фойдаланувчи қўшилди",
                    "username": "system",
                    "created_at": current_timestamp(),
                },
            )
        elif not admin_user["password_hash"]:
            conn.execute(
                text("UPDATE users SET password_hash = :password_hash WHERE username = :username"),
                {"username": "admin", "password_hash": generate_password_hash("admin123")},
            )
            conn.execute(
                text(
                    """
                    INSERT INTO audit_logs(id, action, details, username, created_at)
                    VALUES(:id, :action, :details, :username, :created_at)
                    """
                ),
                {
                    "id": next_id(conn, "audit_logs"),
                    "action": "Admin пароль тикланди",
                    "details": "admin123 вақтинча пароль ўрнатилди",
                    "username": "system",
                    "created_at": current_timestamp(),
                },
            )


def ensure_schema_compatibility(conn) -> None:
    """Add missing columns when an existing Railway/Postgres DB is reused."""
    columns_by_table = {
        "users": {
            "username": "TEXT",
            "password_hash": "TEXT",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "employees": {
            "full_name": "TEXT",
            "is_active": "INTEGER DEFAULT 1",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "operators": {
            "name": "TEXT",
            "is_active": "INTEGER DEFAULT 1",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "tariffs": {
            "operator_id": "INTEGER",
            "name": "TEXT",
            "bonus_per_item": "INTEGER DEFAULT 0",
            "is_active": "INTEGER DEFAULT 1",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "sales": {
            "sale_date": "TEXT",
            "employee_id": "INTEGER",
            "tariff_id": "INTEGER",
            "quantity": "INTEGER DEFAULT 0",
            "comment": "TEXT",
            "created_by": "TEXT",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
            "updated_at": "TIMESTAMP",
        },
        "month_locks": {
            "closed_by": "TEXT",
            "closed_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
        "audit_logs": {
            "action": "TEXT",
            "details": "TEXT",
            "username": "TEXT",
            "created_at": "TIMESTAMP DEFAULT CURRENT_TIMESTAMP",
        },
    }

    if engine.dialect.name == "postgresql":
        for table, columns in columns_by_table.items():
            for column, column_type in columns.items():
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {column} {column_type}"))

        postgres_defaults = {
            "users": {"created_at": "CURRENT_TIMESTAMP"},
            "employees": {"is_active": "1", "created_at": "CURRENT_TIMESTAMP"},
            "operators": {"is_active": "1", "created_at": "CURRENT_TIMESTAMP"},
            "tariffs": {"bonus_per_item": "0", "is_active": "1", "created_at": "CURRENT_TIMESTAMP"},
            "sales": {"quantity": "0", "created_at": "CURRENT_TIMESTAMP"},
            "month_locks": {"closed_at": "CURRENT_TIMESTAMP"},
            "audit_logs": {"created_at": "CURRENT_TIMESTAMP"},
        }
        for table, columns in postgres_defaults.items():
            for column, default_expr in columns.items():
                conn.execute(text(f"ALTER TABLE {table} ALTER COLUMN {column} SET DEFAULT {default_expr}"))
        return

    # SQLite учун: ALTER TABLE ... ADD COLUMN IF NOT EXISTS ҳамма муҳитда ишламаслиги мумкин.
    # Бундан ташқари SQLite мавжуд жадвалга CURRENT_TIMESTAMP default билан устун қўшишга рухсат бермаслиги мумкин.
    for table, columns in columns_by_table.items():
        existing = {row[1] for row in conn.execute(text(f"PRAGMA table_info({table})")).fetchall()}
        for column, column_type in columns.items():
            if column not in existing:
                safe_column_type = column_type.replace(" DEFAULT CURRENT_TIMESTAMP", "")
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {safe_column_type}"))



def ensure_default_operators(conn) -> None:
    """Create default operators and attach all old tariffs/sales to Mobiuz."""
    mobiuz = fetch_one(
        conn,
        "SELECT id FROM operators WHERE lower(trim(name)) = lower(trim(:name))",
        {"name": "Mobiuz"},
    )
    if not mobiuz:
        mobiuz_id = next_id(conn, "operators")
        conn.execute(
            text(
                """
                INSERT INTO operators(id, name, is_active, created_at)
                VALUES(:id, :name, 1, :created_at)
                """
            ),
            {"id": mobiuz_id, "name": "Mobiuz", "created_at": current_timestamp()},
        )
    else:
        mobiuz_id = int(mobiuz["id"])

    uzmobile = fetch_one(
        conn,
        "SELECT id FROM operators WHERE lower(trim(name)) = lower(trim(:name))",
        {"name": "Uzmobile"},
    )
    if not uzmobile:
        conn.execute(
            text(
                """
                INSERT INTO operators(id, name, is_active, created_at)
                VALUES(:id, :name, 1, :created_at)
                """
            ),
            {"id": next_id(conn, "operators"), "name": "Uzmobile", "created_at": current_timestamp()},
        )

    conn.execute(
        text("UPDATE tariffs SET operator_id = :mobiuz_id WHERE operator_id IS NULL OR operator_id = 0"),
        {"mobiuz_id": mobiuz_id},
    )

    if engine.dialect.name == "postgresql":
        constraints = conn.execute(
            text(
                """
                SELECT c.conname
                FROM pg_constraint c
                JOIN pg_class t ON t.oid = c.conrelid
                WHERE t.relname = 'tariffs'
                  AND c.contype = 'u'
                  AND pg_get_constraintdef(c.oid) IN ('UNIQUE (name)', 'UNIQUE(name)')
                """
            )
        ).fetchall()
        for row in constraints:
            constraint_name = str(row[0]).replace('"', '""')
            conn.execute(text(f'ALTER TABLE tariffs DROP CONSTRAINT IF EXISTS "{constraint_name}"'))

    conn.execute(
        text("CREATE UNIQUE INDEX IF NOT EXISTS uq_tariffs_operator_name ON tariffs(operator_id, name)")
    )


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return func(*args, **kwargs)

    return wrapper


def parse_int(value: Any) -> int:
    try:
        if value is None:
            return 0
        return int(str(value).replace(" ", "").replace(",", ""))
    except (TypeError, ValueError):
        return 0


def money(value: Any) -> str:
    try:
        return f"{int(value):,}".replace(",", " ")
    except Exception:
        return "0"


def qty(value: Any) -> str:
    try:
        return f"{int(value):,}".replace(",", " ")
    except Exception:
        return "0"


def tariff_name_display(value: Any) -> str:
    """Тариф номида пул бирлиги ёзилган бўлса, экранда дона бирлигида кўрсатади."""
    name = str(value or "").strip()
    if not name:
        return ""
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r"(?iu)\b(сум|сўм|so['’‘`]?m|sum)\b\.?", "дона", name)
    if re.fullmatch(r"\d+(?:[\s.,]\d+)*", name):
        return f"{name} дона"
    return name


def is_valid_date(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def is_valid_month(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m")
        return True
    except ValueError:
        return False


def validate_sale_form(sale_date: str, employee_id: int, tariff_id: int, quantity: int) -> str | None:
    if not is_valid_date(sale_date):
        return "Сана нотўғри киритилган."
    if employee_id <= 0:
        return "Ходим танланмаган."
    if tariff_id <= 0:
        return "Тариф танланмаган."
    if quantity <= 0:
        return "Сотилган SIM-карта сони 0 дан катта бўлиши керак."
    return None


def collect_tariff_quantities(form: Any) -> tuple[dict[int, int], bool]:
    """Return tariff_id -> quantity from the convenient daily sales table."""
    quantities: dict[int, int] = {}
    has_negative = False
    for key in form.keys():
        key_text = str(key)
        if not key_text.startswith("quantity_"):
            continue
        tariff_id = parse_int(key_text.replace("quantity_", "", 1))
        quantity = parse_int(form.get(key_text))
        if quantity < 0:
            has_negative = True
        elif tariff_id > 0 and quantity > 0:
            quantities[tariff_id] = quantity
    return quantities, has_negative


def validate_daily_sale_form(
    sale_date: str,
    employee_id: int,
    quantities: dict[int, int],
    has_negative: bool = False,
) -> str | None:
    if not is_valid_date(sale_date):
        return "Сана нотўғри киритилган."
    if employee_id <= 0:
        return "Ходим танланмаган."
    if has_negative:
        return "Сотилган SIM-карта сони манфий бўлмаслиги керак."
    if not quantities:
        return "Камида битта тариф бўйича сотилган сонини киритинг."
    return None


def save_daily_sales(
    sale_date: str,
    employee_id: int,
    quantities: dict[int, int],
    comment: str,
    created_by: str,
) -> tuple[int, int]:
    """Insert/update one employee's daily sales for all entered tariffs."""
    inserted = 0
    updated = 0
    with engine.begin() as conn:
        employee = fetch_one(
            conn,
            "SELECT id FROM employees WHERE id = :id AND is_active = 1",
            {"id": employee_id},
        )
        if not employee:
            return 0, 0

        tariff_rows = fetch_all(
            conn,
            """
            SELECT t.id, t.name
            FROM tariffs t
            JOIN operators o ON o.id = t.operator_id
            WHERE t.is_active = 1 AND o.is_active = 1 AND COALESCE(t.bonus_per_item, 0) > 0
            ORDER BY o.name, t.name
            """,
        )
        active_tariffs = {int(row["id"]): row["name"] for row in tariff_rows}

        for tariff_id, quantity in quantities.items():
            if tariff_id not in active_tariffs or quantity <= 0:
                continue
            existing = fetch_one(
                conn,
                """
                SELECT id FROM sales
                WHERE sale_date = :sale_date AND employee_id = :employee_id AND tariff_id = :tariff_id
                """,
                {"sale_date": sale_date, "employee_id": employee_id, "tariff_id": tariff_id},
            )
            if existing:
                conn.execute(
                    text(
                        """
                        UPDATE sales
                        SET quantity = :quantity,
                            comment = :comment,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = :id
                        """
                    ),
                    {"quantity": quantity, "comment": comment, "id": existing["id"]},
                )
                add_audit(
                    conn,
                    "Кунлик сотув янгиланди",
                    f"{sale_date}, employee_id={employee_id}, tariff_id={tariff_id}, quantity={quantity}",
                )
                updated += 1
            else:
                conn.execute(
                    text(
                        """
                        INSERT INTO sales(id, sale_date, employee_id, tariff_id, quantity, comment, created_by, created_at)
                        VALUES(:id, :sale_date, :employee_id, :tariff_id, :quantity, :comment, :created_by, :created_at)
                        """
                    ),
                    {
                        "id": next_id(conn, "sales"),
                        "sale_date": sale_date,
                        "employee_id": employee_id,
                        "tariff_id": tariff_id,
                        "quantity": quantity,
                        "comment": comment,
                        "created_by": created_by,
                        "created_at": current_timestamp(),
                    },
                )
                add_audit(
                    conn,
                    "Сотув қўшилди",
                    f"{sale_date}, employee_id={employee_id}, tariff_id={tariff_id}, quantity={quantity}",
                )
                inserted += 1
    return inserted, updated


def get_active_refs(
    include_ids: tuple[int, int] | None = None,
    positive_tariffs_only: bool = False,
) -> tuple[list[Any], list[Any]]:
    include_employee_id, include_tariff_id = include_ids or (0, 0)
    if positive_tariffs_only:
        tariff_sql = """
            SELECT t.*, o.name AS operator_name
            FROM tariffs t
            JOIN operators o ON o.id = t.operator_id
            WHERE ((t.is_active = 1 AND o.is_active = 1 AND COALESCE(t.bonus_per_item, 0) > 0) OR t.id = :id)
            ORDER BY o.name, t.name
        """
    else:
        tariff_sql = """
            SELECT t.*, o.name AS operator_name
            FROM tariffs t
            JOIN operators o ON o.id = t.operator_id
            WHERE t.is_active = 1 OR t.id = :id
            ORDER BY o.name, t.name
        """
    with engine.connect() as conn:
        employees = fetch_all(
            conn,
            "SELECT * FROM employees WHERE is_active = 1 OR id = :id ORDER BY full_name",
            {"id": include_employee_id},
        )
        tariffs = fetch_all(conn, tariff_sql, {"id": include_tariff_id})
    return employees, tariffs


def get_all_refs() -> tuple[list[Any], list[Any], list[Any]]:
    with engine.connect() as conn:
        employees = fetch_all(conn, "SELECT * FROM employees ORDER BY full_name")
        tariffs = fetch_all(
            conn,
            """
            SELECT t.*, o.name AS operator_name
            FROM tariffs t
            JOIN operators o ON o.id = t.operator_id
            ORDER BY o.name, t.name
            """,
        )
        operators = fetch_all(conn, "SELECT * FROM operators ORDER BY is_active DESC, name")
    return employees, tariffs, operators


def get_filters() -> dict[str, Any]:
    return {
        "date_from": request.args.get("date_from", "").strip(),
        "date_to": request.args.get("date_to", "").strip(),
        "employee_id": parse_int(request.args.get("employee_id")),
        "operator_id": parse_int(request.args.get("operator_id")),
        "tariff_id": parse_int(request.args.get("tariff_id")),
        "q": request.args.get("q", "").strip(),
    }


def fetch_sales(filters: dict[str, Any]) -> list[Any]:
    sql = [
        """
        SELECT s.*, e.full_name, t.name AS tariff_name, t.bonus_per_item,
               o.id AS operator_id, o.name AS operator_name,
               (s.quantity * t.bonus_per_item) AS bonus_total
        FROM sales s
        JOIN employees e ON e.id = s.employee_id
        JOIN tariffs t ON t.id = s.tariff_id
        JOIN operators o ON o.id = t.operator_id
        WHERE 1=1
        """
    ]
    params: dict[str, Any] = {}
    if filters["date_from"]:
        sql.append("AND s.sale_date >= :date_from")
        params["date_from"] = filters["date_from"]
    if filters["date_to"]:
        sql.append("AND s.sale_date <= :date_to")
        params["date_to"] = filters["date_to"]
    if filters["employee_id"]:
        sql.append("AND s.employee_id = :employee_id")
        params["employee_id"] = filters["employee_id"]
    if filters["operator_id"]:
        sql.append("AND t.operator_id = :operator_id")
        params["operator_id"] = filters["operator_id"]
    if filters["tariff_id"]:
        sql.append("AND s.tariff_id = :tariff_id")
        params["tariff_id"] = filters["tariff_id"]
    if filters["q"]:
        sql.append("AND (e.full_name LIKE :q OR o.name LIKE :q OR t.name LIKE :q OR s.comment LIKE :q)")
        params["q"] = f"%{filters['q']}%"
    sql.append("ORDER BY s.sale_date DESC, o.name, e.full_name, t.name")
    with engine.connect() as conn:
        return fetch_all(conn, " ".join(sql), params)


def build_report(month: str, operator_id: int = 0) -> dict[str, Any]:
    if not is_valid_month(month):
        month = date.today().strftime("%Y-%m")
    operator_where = " AND t.operator_id = :operator_id" if operator_id else ""
    params = {"month": month, "operator_id": operator_id}
    with engine.connect() as conn:
        detail = fetch_all(
            conn,
            f"""
            SELECT o.name AS operator_name, e.full_name, t.name AS tariff_name, t.bonus_per_item,
                   SUM(s.quantity) AS total_qty,
                   SUM(s.quantity * t.bonus_per_item) AS total_bonus
            FROM sales s
            JOIN employees e ON e.id = s.employee_id
            JOIN tariffs t ON t.id = s.tariff_id
            JOIN operators o ON o.id = t.operator_id
            WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
            GROUP BY o.id, o.name, e.id, e.full_name, t.id, t.name, t.bonus_per_item
            ORDER BY o.name, e.full_name, t.name
            """,
            params,
        )
        summary = fetch_all(
            conn,
            f"""
            SELECT e.full_name,
                   SUM(s.quantity) AS total_qty,
                   SUM(s.quantity * t.bonus_per_item) AS total_bonus
            FROM sales s
            JOIN employees e ON e.id = s.employee_id
            JOIN tariffs t ON t.id = s.tariff_id
            WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
            GROUP BY e.id, e.full_name
            ORDER BY total_bonus DESC, e.full_name
            """,
            params,
        )
        operator_summary = fetch_all(
            conn,
            f"""
            SELECT o.id AS operator_id, o.name AS operator_name,
                   SUM(s.quantity) AS total_qty,
                   SUM(s.quantity * t.bonus_per_item) AS total_bonus
            FROM sales s
            JOIN tariffs t ON t.id = s.tariff_id
            JOIN operators o ON o.id = t.operator_id
            WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
            GROUP BY o.id, o.name
            ORDER BY o.name
            """,
            params,
        )
        totals = fetch_one(
            conn,
            f"""
            SELECT COALESCE(SUM(s.quantity), 0) AS total_qty,
                   COALESCE(SUM(s.quantity * t.bonus_per_item), 0) AS total_bonus
            FROM sales s
            JOIN tariffs t ON t.id = s.tariff_id
            WHERE substr(s.sale_date, 1, 7) = :month {operator_where}
            """,
            params,
        )
    return {
        "detail": detail,
        "summary": summary,
        "operator_summary": operator_summary,
        "totals": totals or {"total_qty": 0, "total_bonus": 0},
    }


def is_month_closed(month: str) -> bool:
    if not is_valid_month(month):
        return False
    with engine.connect() as conn:
        row = fetch_one(conn, "SELECT month FROM month_locks WHERE month = :month", {"month": month})
    return row is not None


def add_audit(conn, action: str, details: str) -> None:
    username = session.get("username", "admin") if session else "admin"
    conn.execute(
        text(
            """
            INSERT INTO audit_logs(id, action, details, username, created_at)
            VALUES(:id, :action, :details, :username, :created_at)
            """
        ),
        {
            "id": next_id(conn, "audit_logs"),
            "action": action,
            "details": details,
            "username": username,
            "created_at": current_timestamp(),
        },
    )


def create_excel_report(file_path: Path, month: str, data: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ойлик ҳисобот"
    ws.append([f"I-MAX — SIM-карта сотувлари бўйича бонус ҳисоботи — {month}"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(["Оператор", "Ходим Ф.И.О.", "Тариф номи", "Сони", "1 дона бонус", "Жами бонус"])
    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    for row in data["detail"]:
        ws.append([
            row["operator_name"],
            row["full_name"],
            tariff_name_display(row["tariff_name"]),
            row["total_qty"],
            row["bonus_per_item"],
            row["total_bonus"],
        ])
    ws.append([])
    ws.append(["ЖАМИ", "", "", data["totals"]["total_qty"], "", data["totals"]["total_bonus"]])
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("Ходимлар бўйича жами")
    ws2.append([f"I-MAX — ходимлар бўйича умумий натижа — {month}"])
    ws2.merge_cells("A1:C1")
    ws2["A1"].font = Font(size=14, bold=True)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.append([])
    ws2.append(["Ходим Ф.И.О.", "Жами SIM", "Жами бонус"])
    for cell in ws2[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for row in data["summary"]:
        ws2.append([row["full_name"], row["total_qty"], row["total_bonus"]])
    ws2.append([])
    ws2.append(["ЖАМИ", data["totals"]["total_qty"], data["totals"]["total_bonus"]])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)

    ws3 = wb.create_sheet("Операторлар бўйича")
    ws3.append([f"I-MAX — операторлар бўйича натижа — {month}"])
    ws3.merge_cells("A1:C1")
    ws3["A1"].font = Font(size=14, bold=True)
    ws3["A1"].alignment = Alignment(horizontal="center")
    ws3.append([])
    ws3.append(["Оператор", "Жами SIM", "Жами бонус"])
    for cell in ws3[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for row in data["operator_summary"]:
        ws3.append([row["operator_name"], row["total_qty"], row["total_bonus"]])
    ws3.append([])
    ws3.append(["ЖАМИ", data["totals"]["total_qty"], data["totals"]["total_bonus"]])
    for cell in ws3[ws3.max_row]:
        cell.font = Font(bold=True)

    thin = Side(style="thin", color="CCCCCC")
    for sheet in (ws, ws2, ws3):
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col_idx in range(1, sheet.max_column + 1):
            max_len = 12
            col_letter = get_column_letter(col_idx)
            for cell in sheet[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)) + 2)
            sheet.column_dimensions[col_letter].width = min(max_len, 45)
        for row_idx in range(4, sheet.max_row + 1):
            for col_idx in range(2, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                if isinstance(cell.value, int):
                    cell.number_format = '#,##0'

    wb.save(file_path)


app = create_app()
app.jinja_env.filters["money"] = money
app.jinja_env.filters["qty"] = qty
app.jinja_env.filters["tariff_name"] = tariff_name_display

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(debug=os.getenv("FLASK_DEBUG") == "1", host="0.0.0.0", port=port)
